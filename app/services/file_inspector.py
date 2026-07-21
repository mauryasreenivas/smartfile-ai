from pathlib import Path
from typing import Any

import polars as pl
from openpyxl import load_workbook

from app.core.exceptions import InvalidFileError
from app.schemas.file import (
    CsvInspection,
    ExcelInspection,
    ExcelSheetInspection,
)


class FileInspectorService:
    def __init__(self, preview_rows: int = 10) -> None:
        self.preview_rows = preview_rows

    def inspect(self, file_path: Path) -> CsvInspection | ExcelInspection:
        extension = file_path.suffix.lower()

        if extension == ".csv":
            return self._inspect_csv(file_path)

        if extension == ".xlsx":
            return self._inspect_excel(file_path)

        raise InvalidFileError(f"No inspector exists for extension: {extension}")

    def _inspect_csv(self, file_path: Path) -> CsvInspection:
        try:
            dataframe = pl.read_csv(
                file_path,
                infer_schema_length=1000,
                ignore_errors=False,
                try_parse_dates=False,
            )

            preview = dataframe.head(self.preview_rows).to_dicts()

            return CsvInspection(
                row_count=dataframe.height,
                column_count=dataframe.width,
                columns=dataframe.columns,
                preview=self._normalise_records(preview),
            )

        except Exception as exc:
            raise InvalidFileError(f"Unable to read CSV file: {exc}") from exc

    def _inspect_excel(self, file_path: Path) -> ExcelInspection:
        try:
            workbook = load_workbook(
                filename=file_path,
                read_only=True,
                data_only=True,
            )

            sheet_inspections: list[ExcelSheetInspection] = []

            try:
                for worksheet in workbook.worksheets:
                    rows = worksheet.iter_rows(values_only=True)

                    header_row = next(rows, None)

                    if header_row is None:
                        sheet_inspections.append(
                            ExcelSheetInspection(
                                sheet_name=worksheet.title,
                                estimated_row_count=0,
                                estimated_column_count=0,
                                columns=[],
                                preview=[],
                            )
                        )
                        continue

                    columns = self._make_unique_headers(header_row)
                    preview: list[dict[str, Any]] = []

                    for index, row in enumerate(rows):
                        if index >= self.preview_rows:
                            break

                        padded_row = list(row) + [None] * (len(columns) - len(row))

                        record = dict(zip(columns, padded_row, strict=False))
                        preview.append(record)

                    estimated_rows = max(worksheet.max_row - 1, 0)

                    sheet_inspections.append(
                        ExcelSheetInspection(
                            sheet_name=worksheet.title,
                            estimated_row_count=estimated_rows,
                            estimated_column_count=len(columns),
                            columns=columns,
                            preview=self._normalise_records(preview),
                        )
                    )

            finally:
                workbook.close()

            return ExcelInspection(
                sheet_count=len(sheet_inspections),
                sheets=sheet_inspections,
            )

        except Exception as exc:
            raise InvalidFileError(f"Unable to read Excel file: {exc}") from exc

    @staticmethod
    def _make_unique_headers(values: tuple[Any, ...]) -> list[str]:
        headers: list[str] = []
        counts: dict[str, int] = {}

        for index, value in enumerate(values, start=1):
            base_header = (
                str(value).strip()
                if value is not None and str(value).strip()
                else f"column_{index}"
            )

            count = counts.get(base_header, 0)
            counts[base_header] = count + 1

            if count == 0:
                headers.append(base_header)
            else:
                headers.append(f"{base_header}_{count + 1}")

        return headers

    @classmethod
    def _normalise_records(
        cls,
        records: list[dict[str, Any]],
    ) -> list[dict[str, Any]]:
        return [
            {key: cls._normalise_value(value) for key, value in record.items()}
            for record in records
        ]

    @staticmethod
    def _normalise_value(value: Any) -> Any:
        if value is None:
            return None

        if hasattr(value, "isoformat"):
            return value.isoformat()

        return value
